import argparse
import sys
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from typing import Optional
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet


# ---------------------------------------------------------------------------
# COLAB SETTINGS
# ---------------------------------------------------------------------------

# Change these only if needed
WIP_FILE = "/content/Narrow Data_Coil Stage.xlsx"
PLAN_DATE = "2026-05-21"
OUTPUT_FILE = "/content/mill_plan_output.xlsx"
DAYS = 3


# ---------------------------------------------------------------------------
# CONSTANTS
# ---------------------------------------------------------------------------

EXCEL_EPOCH = date(1899, 12, 30)

REQUIRED_COLUMNS = [
    'Planning Date',
    'Coil Number',
    'SO No',
    'Actual Thick',
    'Actual Width',
    'Input Coil Weight',
    'Plan Rolling Thick 1',
    'Customer Desc',
    'Product Code',
    'Actual Quality',
    'Cust TDC',
    'Production Plant',
    'Storage Location',
    'Planning Remark',
    'Current Stage',
    'Next Stage',
    'Process Route',
    'Edge',
    'Coil Age(# Days)',
    'Surface Finish',
    'Work Center',
    'Cust Thick',
    'Plan Weight',
    'Balance Coil Weight',
    'Last Production Stage'
]

PLAN_COLUMNS = [
    "Date", "Batch", "Planning/ SO No.", "Thick", "Width", "Weight", "RT",
    "Customer", "Prod.Code", "Quality Code", "TDC No", "Plant", "Storage Loc",
    "Planning Remark", "Current Work Center", "NEXT Work Center", "Route",
    "Finish", "Age",
]

COLUMN_WIDTHS = {
    "A": 8,  "B": 12, "C": 16, "D": 7,  "E": 7,  "F": 9,  "G": 7,
    "H": 14, "I": 10, "J": 12, "K": 8,  "L": 6,  "M": 10, "N": 22,
    "O": 18, "P": 16, "Q": 32, "R": 10, "S": 6,
}

SECTION_ORDER = [
    "CRCA_FINISH_CRM04",
    "HT_FINISH_CRM04",
    "FIRST_ROLLING_CRM06",
    "RR_ROLLING_CRM06",
    "SKIN_PASS_SUPER_BRIGHT_CRM04",
    "SKIN_PASS_CHROME_CRM04",
    "TUBE_FH",
    "SKIN_PASS_HEAVY_MATT_CRM06",
    "OTHER",
]

SECTION_META = {
    "CRCA_FINISH_CRM04": (
        "CRCA FINISH ON BRIGHT ROLLS --------------- AT CRM-04 ----------- APPLY R.P.OIL",
        "DCE6F1",
    ),
    "HT_FINISH_CRM04": (
        "H&T FINISH ON BRIGHT ROLLS --------------- AT CRM-04 ------------ DO NOT APPLY R.P.OIL",
        "DCE6F1",
    ),
    "FIRST_ROLLING_CRM06": (
        "1ST ROLLING ON LIGHT MATT ROLLS --------------- AT CRM-06",
        "EBF1DE",
    ),
    "RR_ROLLING_CRM06": (
        "R/R ROLLING ON LIGHT MATT ROLLS --------------- AT CRM-06",
        "EBF1DE",
    ),
    "SKIN_PASS_SUPER_BRIGHT_CRM04": (
        "SKIN-PASS ON SUPER BRIGHT ROLLS --------------- AT CRM-04 ----------- APPLY R.P.OIL",
        "FFFF99",
    ),
    "SKIN_PASS_CHROME_CRM04": (
        "SKIN-PASS ON CHROMEPLATED ROLLS --------------- AT CRM-04 ----------- APPLY R.P.OIL",
        "FFFF99",
    ),
    "TUBE_FH": (
        "TUBE FH ON BRIGHT ROLLS --------------- AT CRM-04/06 ----------- APPLY R.P.OIL",
        "FCE4D6",
    ),
    "SKIN_PASS_HEAVY_MATT_CRM06": (
        "SKIN-PASS ON HEAVY MATT ROLLS --------------- AT CRM-06 ----------- APPLY R.P.OIL",
        "FFFF99",
    ),
    "OTHER": (
        "OTHER / CHECK ROUTING --------------- MANUAL REVIEW REQUIRED",
        "D9D9D9",
    ),
}

CUSTOMER_ABBREV = {
    "L.G BALAKRISHNAN": "LG BALA",
    "L.G. BALAKRISHNAN": "LG BALA",
    "SAHIBABAD TUBE PLANT": "TUBE",
    "BIJOY TRADING": "BIJOY",
    "SPECIAL STEEL": "SPECIAL STEEL",
    "BANDSAW STRIP": "BANDSAW",
    "ANCHOR": "ANCHOR",
    "INSAFE SAFETY": "INSAFE",
    "KARAM": "KARAM",
    "MUNJAL": "MUNJAL",
    "CALLIDA": "CALLIDA",
    "VAISH": "VAISH",
    "BOX": "BOX",
}

ROUTE_DISPLAY = {
    "M->RW->B->S->QA->R->QA->PACK": "M->RW->B->S->QA->R->QA->PACK",
    "M->B->M->R->QA->PACK": "M->B->M->R->QA->PACK",
}

CRM04_WORK_CENTERS = {
    "SNCRMM04", "SNCRS13", "SNCRS14", "SNCRS11", "SNCRS10", "SWCRS1"
}

CRM06_WORK_CENTERS = {
    "SNCRMM06", "SNCRS15", "SNRWL06", "SNANN02"
}


# ---------------------------------------------------------------------------
# DATA LOADING & FILTERING
# ---------------------------------------------------------------------------

def validate_required_columns(df: pd.DataFrame) -> None:
    missing = [c for c in REQUIRED_COLUMNS if c not in df.columns]
    if missing:
        raise ValueError(
            "Missing required columns in input file:\n"
            + "\n".join(f"- {c}" for c in missing)
        )


def load_wip(filepath: str) -> pd.DataFrame:
    """Load Sheet1 of the WIP file and normalise text/numeric columns."""
    filepath = str(filepath)

    if not Path(filepath).exists():
        raise FileNotFoundError(filepath)

    df = pd.read_excel(filepath, sheet_name="Sheet1", engine="openpyxl")

    validate_required_columns(df)

    text_dtypes = df.select_dtypes(include=["object", "string"]).columns
    for col in text_dtypes:
        df[col] = df[col].astype(str).str.strip()
        df[col] = df[col].replace({"nan": "", "None": "", "NaT": ""})

    numeric_cols = [
        "Actual Thick",
        "Actual Width",
        "Input Coil Weight",
        "Plan Rolling Thick 1",
        "Cust Width",
        "Cust Thick",
        "Plan Weight",
        "Balance Coil Weight",
        "Coil Age(# Days)",
        "Stage Age(# Days)",
    ]

    for col in numeric_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    return df


def filter_rolling_coils(df: pd.DataFrame, plan_date: date) -> tuple[pd.DataFrame, dict]:
    """
    Eligibility filters:
    1. Current Stage == ROLLING MILL OR Last Production Stage == ROLLING MILL
    2. Exclude FG / palletization / pack / hold
    3. Input Coil Weight >= 0.5
    4. Plan Rolling Thick 1 > 0
    5. Production Plant == 0760
    6. Planning Remark does not contain hold
    """
    counts = {
        "low_weight": 0,
        "fg_or_pack": 0,
        "rt_zero_or_hold": 0,
        "wrong_plant": 0,
        "not_rolling": 0,
        "remark_hold": 0,
    }

    df = df.copy()

    plant = df["Production Plant"].astype(str).str.strip()
    mask_plant = plant.isin(["0760", "760"]) | plant.str.contains("760", na=False)
    counts["wrong_plant"] = int((~mask_plant).sum())
    df = df[mask_plant].copy()

    current_stage = df["Current Stage"].astype(str).str.upper()
    last_stage = df["Last Production Stage"].astype(str).str.upper()

    mask_rolling = (
        current_stage.eq("ROLLING MILL")
        | last_stage.eq("ROLLING MILL")
        | current_stage.str.contains("ROLLING MILL", na=False)
        | last_stage.str.contains("ROLLING MILL", na=False)
    )

    counts["not_rolling"] = int((~mask_rolling).sum())
    df = df[mask_rolling].copy()

    next_stage = df["Next Stage"].astype(str).str.upper().str.strip()

    mask_done = (
        next_stage.isin(["11-FG", "PALLETIZATION", "FG/PALLETIZATION", "PACK", "HOLD"])
        | next_stage.str.contains("PALLET", na=False)
        | next_stage.str.contains("PACK", na=False)
    )

    counts["fg_or_pack"] = int(mask_done.sum())
    df = df[~mask_done].copy()

    remark = df["Planning Remark"].astype(str)
    mask_remark_hold = remark.str.contains("hold", case=False, na=False)
    counts["remark_hold"] = int(mask_remark_hold.sum())
    df = df[~mask_remark_hold].copy()

    mask_low = df["Input Coil Weight"].fillna(0) < 0.5
    counts["low_weight"] = int(mask_low.sum())
    df = df[~mask_low].copy()

    rt = df["Plan Rolling Thick 1"].fillna(0)
    mask_rt_zero = rt <= 0
    counts["rt_zero_or_hold"] = int(mask_rt_zero.sum())
    df = df[~mask_rt_zero].copy()

    return df.copy(), counts


# ---------------------------------------------------------------------------
# SECTION ASSIGNMENT
# ---------------------------------------------------------------------------

def _is_crm04(wc: str) -> bool:
    wc = str(wc or "").upper()
    return any(code in wc for code in CRM04_WORK_CENTERS) or "04" in wc


def _is_crm06(wc: str) -> bool:
    wc = str(wc or "").upper()
    return any(code in wc for code in CRM06_WORK_CENTERS) or "06" in wc


def assign_section(row: pd.Series) -> str:
    """
    Map one WIP row to a plan section key.
    Order matters: specific routing first, generic fallback last.
    """
    wc = str(row.get("Work Center", "")).upper()
    quality = str(row.get("Actual Quality", "")).upper()
    prod_code = str(row.get("Product Code", "")).upper()
    route = str(row.get("Process Route", "")).upper()
    storage = str(row.get("Storage Location", "")).upper()
    next_stage = str(row.get("Next Stage", "")).upper()
    tdc = str(row.get("Cust TDC", "")).upper()
    customer = str(row.get("Customer Desc", "")).upper()
    surface = str(row.get("Surface Finish", "")).upper()
    remark = str(row.get("Planning Remark", "")).upper()

    rt = row.get("Plan Rolling Thick 1") or 0
    actual_thick = row.get("Actual Thick") or 0

    try:
        rt = float(rt)
    except Exception:
        rt = 0.0

    try:
        actual_thick = float(actual_thick)
    except Exception:
        actual_thick = 0.0

    # 1. TUBE FH
    if (
        quality == "TATFHC"
        or "TUBE" in customer
        or "SAHIBABAD TUBE" in customer
    ):
        return "TUBE_FH"

    # 2. HEAVY MATT
    if (
        quality == "TATBID"
        or "S-SPM" in next_stage
        or ("HEAVY" in surface and "MATT" in surface)
    ):
        if _is_crm06(wc) or storage in ("RNM6", "RC01") or "S-SPM" in next_stage:
            return "SKIN_PASS_HEAVY_MATT_CRM06"

    # 3. SKIN PASS CRM04 — D012 chrome / super bright distinction
    skin_pass_route = (
        "S->QA" in route
        or "R->QA->PACK" in route
        or "SKIN" in next_stage
        or "SPM" in next_stage
    )

    if skin_pass_route or prod_code == "C01":
        if "D012" in tdc and rt < 2.0:
            return "SKIN_PASS_CHROME_CRM04"

        if (
            quality in ("TATXXD", "TATD12")
            or "T012" in tdc
            or "SUPER" in surface
            or "BRIGHT" in surface
        ):
            return "SKIN_PASS_SUPER_BRIGHT_CRM04"

    # 4. CRCA FINISH CRM04
    if (
        "CRCA" in route
        or "FINISH" in next_stage
        or (prod_code == "C09" and ("TSBF" in quality or "TSBH" in quality))
    ):
        if _is_crm04(wc) or storage in ("R034", "R032", "R116"):
            return "CRCA_FINISH_CRM04"

    # 5. H&T FINISH CRM04
    if (
        prod_code == "B28"
        or "H&T" in route
        or "HT" in route
        or "HARDEN" in next_stage
    ):
        return "HT_FINISH_CRM04"

    # 6. FIRST ROLLING CRM06
    if (_is_crm06(wc) or storage == "R116") and (
        "B-ANNEALING" in next_stage
        or "ANNEAL" in next_stage
    ):
        return "FIRST_ROLLING_CRM06"

    # 7. R/R ROLLING CRM06
    if (_is_crm06(wc) or storage == "RNM6") and (
        "RW-REWINDING" in next_stage
        or "REWIND" in next_stage
    ):
        return "RR_ROLLING_CRM06"

    # 8. Generic CRM06 rolling fallback
    if _is_crm06(wc):
        return "FIRST_ROLLING_CRM06"

    # 9. Generic CRM04 CRCA/HT fallback
    if _is_crm04(wc):
        if actual_thick >= 2.0:
            return "HT_FINISH_CRM04"
        return "CRCA_FINISH_CRM04"

    return "OTHER"


def assign_sections(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df["section"] = df.apply(assign_section, axis=1)
    return df


def sort_within_section(section_df: pd.DataFrame) -> pd.DataFrame:
    """
    Planner-style sorting:
    1. Width descending
    2. Thickness descending
    3. SO grouping
    4. Age descending
    5. Weight descending
    """
    return section_df.sort_values(
        by=[
            "Actual Width",
            "Actual Thick",
            "SO No",
            "Coil Age(# Days)",
            "Input Coil Weight",
        ],
        ascending=[False, False, True, False, False],
        kind="mergesort",
    )


# ---------------------------------------------------------------------------
# ROW BUILDING
# ---------------------------------------------------------------------------

def _abbreviate_customer(name: str) -> str:
    if not name:
        return ""

    upper = str(name).upper()

    for key, short in CUSTOMER_ABBREV.items():
        if key.upper() in upper:
            return short

    return str(name)[:12]


def _route_display(route: str) -> str:
    route = str(route or "")
    return ROUTE_DISPLAY.get(route, route)


def _plant_short(plant: str) -> str:
    s = str(plant).strip().lstrip("0")
    return s or "0"


def _excel_serial(d: date) -> int:
    return (d - EXCEL_EPOCH).days


def build_plan_row(wip_row: pd.Series, plan_date: date, age_offset: int = 0) -> list:
    serial = _excel_serial(plan_date)

    age = wip_row.get("Coil Age(# Days)")
    try:
        age_int = int(age) + age_offset if pd.notna(age) else age_offset
    except Exception:
        age_int = age_offset

    finish = str(wip_row.get("Surface Finish") or "").strip()
    if not finish:
        finish = str(wip_row.get("Edge") or "").strip()

    return [
        serial,
        str(wip_row.get("Coil Number", "")),
        str(wip_row.get("SO No", "")),
        float(wip_row.get("Actual Thick") or 0),
        int(wip_row.get("Actual Width") or 0),
        float(wip_row.get("Input Coil Weight") or 0),
        float(wip_row.get("Plan Rolling Thick 1") or 0),
        _abbreviate_customer(str(wip_row.get("Customer Desc", ""))),
        str(wip_row.get("Product Code", "")),
        str(wip_row.get("Actual Quality", "")),
        str(wip_row.get("Cust TDC", "")),
        _plant_short(wip_row.get("Production Plant", "")),
        str(wip_row.get("Storage Location", "")),
        str(wip_row.get("Planning Remark", "")),
        str(wip_row.get("Current Stage", "")),
        str(wip_row.get("Next Stage", "")),
        _route_display(str(wip_row.get("Process Route", ""))),
        finish,
        age_int,
    ]


# ---------------------------------------------------------------------------
# EXCEL STYLES
# ---------------------------------------------------------------------------

THIN = Side(border_style="thin", color="999999")
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

HEADER_FONT = Font(name="Calibri", size=10, bold=True, color="000000")
SECTION_FONT = Font(name="Calibri", size=10, bold=True)
DATA_FONT = Font(name="Calibri", size=9)
SUBTOTAL_FONT = Font(name="Calibri", size=10, bold=True)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")

HEADER_FILL = PatternFill("solid", start_color="D9D9D9", end_color="D9D9D9")
TITLE_FILL = PatternFill("solid", start_color="305496", end_color="305496")
SUBTOTAL_FILL = PatternFill("solid", start_color="F2F2F2", end_color="F2F2F2")
GRAND_TOTAL_FILL = PatternFill("solid", start_color="305496", end_color="305496")


# ---------------------------------------------------------------------------
# EXCEL WRITING
# ---------------------------------------------------------------------------

def _apply_column_widths(ws: Worksheet) -> None:
    for col_letter, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width


def _write_sheet_title(ws: Worksheet, plan_date: date, row: int) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=19)

    cell = ws.cell(row=row, column=1)
    cell.value = f"PLANNING FOR MILL-------------- AS ON  {plan_date.strftime('%d-%m-%Y')}"
    cell.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    cell.fill = TITLE_FILL
    cell.alignment = CENTER

    ws.row_dimensions[row].height = 22
    return row + 1


def _write_header_row(ws: Worksheet, row: int) -> int:
    for col_idx, name in enumerate(PLAN_COLUMNS, start=1):
        cell = ws.cell(row=row, column=col_idx, value=name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = CELL_BORDER

    ws.row_dimensions[row].height = 30
    return row + 1


def _write_section_header(ws: Worksheet, label: str, fill_hex: str, row: int) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=19)

    cell = ws.cell(row=row, column=1, value=label)
    cell.font = SECTION_FONT
    cell.fill = PatternFill("solid", start_color=fill_hex, end_color=fill_hex)
    cell.alignment = LEFT
    cell.border = CELL_BORDER

    # Apply fill/border across merged row cells
    for col_idx in range(1, 20):
        c = ws.cell(row=row, column=col_idx)
        c.fill = PatternFill("solid", start_color=fill_hex, end_color=fill_hex)
        c.border = CELL_BORDER

    ws.row_dimensions[row].height = 18
    return row + 1


def _write_data_row(ws: Worksheet, values: list, row: int) -> int:
    for col_idx, val in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.font = DATA_FONT
        cell.border = CELL_BORDER

        if col_idx == 1:
            cell.number_format = "0"
            cell.alignment = CENTER
        elif col_idx in (4, 7):
            cell.number_format = "0.00"
            cell.alignment = RIGHT
        elif col_idx == 5:
            cell.number_format = "0"
            cell.alignment = RIGHT
        elif col_idx == 6:
            cell.number_format = "0.000"
            cell.alignment = RIGHT
        elif col_idx == 19:
            cell.number_format = "0"
            cell.alignment = CENTER
        else:
            cell.alignment = LEFT

    return row + 1


def _write_subtotal_row(ws: Worksheet, first_data_row: int, last_data_row: int, row: int) -> int:
    if last_data_row >= first_data_row:
        ws.cell(row=row, column=6).value = f"=SUM(F{first_data_row}:F{last_data_row})"
    else:
        ws.cell(row=row, column=6).value = 0

    for col_idx in range(1, 20):
        cell = ws.cell(row=row, column=col_idx)
        cell.font = SUBTOTAL_FONT
        cell.border = CELL_BORDER
        cell.fill = SUBTOTAL_FILL

        if col_idx == 6:
            cell.number_format = "0.000"
            cell.alignment = RIGHT
        elif col_idx == 3:
            cell.value = "SUB-TOTAL"
            cell.alignment = RIGHT

    return row + 1


def _write_grand_total(ws: Worksheet, subtotal_rows: list[int], row: int) -> int:
    formula = "=" + "+".join(f"F{r}" for r in subtotal_rows) if subtotal_rows else "=0"

    for col_idx in range(1, 20):
        cell = ws.cell(row=row, column=col_idx)
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.fill = GRAND_TOTAL_FILL
        cell.border = CELL_BORDER

    ws.cell(row=row, column=3, value="GRAND TOTAL").alignment = RIGHT

    gt = ws.cell(row=row, column=6, value=formula)
    gt.number_format = "0.000"
    gt.alignment = RIGHT

    return row + 1


# ---------------------------------------------------------------------------
# SHEET BUILDER
# ---------------------------------------------------------------------------

@dataclass
class SectionResult:
    key: str
    coils: int
    weight: float


def write_day_sheet(
    wb: Workbook,
    df: pd.DataFrame,
    plan_date: date,
    age_offset: int = 0,
    hide_empty_sections: bool = False,
) -> list[SectionResult]:

    sheet_name = plan_date.strftime("%d-%m-%Y")

    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    ws = wb.create_sheet(sheet_name)

    _apply_column_widths(ws)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    row = _write_sheet_title(ws, plan_date, 1)
    row = _write_header_row(ws, row)

    df = assign_sections(df)

    results: list[SectionResult] = []
    subtotal_rows: list[int] = []

    for section_key in SECTION_ORDER:
        label, fill = SECTION_META[section_key]

        section_df = df[df["section"] == section_key]
        section_df = sort_within_section(section_df)

        if hide_empty_sections and len(section_df) == 0:
            results.append(
                SectionResult(
                    key=section_key,
                    coils=0,
                    weight=0.0,
                )
            )
            continue

        row = _write_section_header(ws, label, fill, row)

        first_data_row = row

        for _, wip_row in section_df.iterrows():
            plan_row_values = build_plan_row(wip_row, plan_date, age_offset)
            row = _write_data_row(ws, plan_row_values, row)

        last_data_row = row - 1

        subtotal_rows.append(row)
        row = _write_subtotal_row(ws, first_data_row, last_data_row, row)

        results.append(
            SectionResult(
                key=section_key,
                coils=len(section_df),
                weight=float(section_df["Input Coil Weight"].sum() or 0),
            )
        )

    _write_grand_total(ws, subtotal_rows, row)

    return results


# ---------------------------------------------------------------------------
# VALIDATION PRINTOUT
# ---------------------------------------------------------------------------

SECTION_LABELS = {
    "CRCA_FINISH_CRM04": "CRCA Finish (CRM04)",
    "HT_FINISH_CRM04": "H&T Finish  (CRM04)",
    "FIRST_ROLLING_CRM06": "1st Rolling (CRM06)",
    "RR_ROLLING_CRM06": "R/R Rolling (CRM06)",
    "SKIN_PASS_SUPER_BRIGHT_CRM04": "Skin Pass Super Bright (CRM04)",
    "SKIN_PASS_CHROME_CRM04": "Skin Pass Chrome       (CRM04)",
    "TUBE_FH": "Tube FH     (CRM04/06)",
    "SKIN_PASS_HEAVY_MATT_CRM06": "Skin Pass Heavy Matt   (CRM06)",
    "OTHER": "Other / Manual Review",
}


def print_validation(
    plan_date: date,
    results: list[SectionResult],
    exclusions: dict,
) -> None:

    total_coils = sum(r.coils for r in results)
    total_weight = sum(r.weight for r in results)

    print(f"\n=== MILL PLAN VALIDATION: {plan_date.strftime('%d-%m-%Y')} ===")
    print(f"Total coils planned:          {total_coils}")
    print(f"Total weight (MT):            {total_weight:,.3f}")
    print()

    for r in results:
        label = SECTION_LABELS.get(r.key, r.key)
        print(f"  {label:<36} {r.coils:>4} coils, {r.weight:>10.3f} MT")

    print()
    print(f"Coils excluded (low weight):   {exclusions.get('low_weight', 0)}")
    print(f"Coils excluded (FG/Pack):      {exclusions.get('fg_or_pack', 0)}")
    print(f"Coils excluded (RT=0/HOLD):    {exclusions.get('rt_zero_or_hold', 0)}")
    print(f"Coils excluded (not rolling):  {exclusions.get('not_rolling', 0)}")
    print(f"Coils excluded (wrong plant):  {exclusions.get('wrong_plant', 0)}")
    print(f"Coils excluded (remark hold):  {exclusions.get('remark_hold', 0)}")


# ---------------------------------------------------------------------------
# ORCHESTRATION
# ---------------------------------------------------------------------------

def generate_daily_plan(
    wip_file: str,
    plan_date: date,
    output_file: str,
    days: int = 1,
    hide_empty_sections: bool = False,
) -> str:

    print(f"Loading WIP file: {wip_file}")

    wip = load_wip(wip_file)
    print(f"Rows loaded: {len(wip)}")

    eligible, exclusions = filter_rolling_coils(wip, plan_date)
    print(f"Rows eligible for rolling: {len(eligible)}")

    if eligible.empty:
        raise ValueError("No eligible rolling coils found after filtering.")

    wb = Workbook()

    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    for day_idx in range(days):
        d = plan_date + timedelta(days=day_idx)

        results = write_day_sheet(
            wb=wb,
            df=eligible,
            plan_date=d,
            age_offset=day_idx,
            hide_empty_sections=hide_empty_sections,
        )

        print_validation(d, results, exclusions)

    output_path = Path(output_file)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb.save(output_file)

    print(f"\nWritten: {output_file}")
    return output_file


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def _parse_date(s: str) -> date:
    return datetime.strptime(s, "%Y-%m-%d").date()


def main(argv: Optional[list[str]] = None) -> int:
    parser = argparse.ArgumentParser(
        description="Generate Tata Steel CRM Narrow Complex daily mill plan."
    )

    parser.add_argument(
        "--wip_file",
        required=True,
        help="Path to Narrow_Data_Coil_Stage.xlsx",
    )

    parser.add_argument(
        "--plan_date",
        required=True,
        type=_parse_date,
        help="Planning date as YYYY-MM-DD",
    )

    parser.add_argument(
        "--output",
        required=True,
        help="Output mill plan .xlsx path",
    )

    parser.add_argument(
        "--days",
        type=int,
        default=1,
        help="Number of consecutive days to generate",
    )

    parser.add_argument(
        "--hide_empty_sections",
        action="store_true",
        help="Hide empty section headers",
    )

    args = parser.parse_args(argv)

    generate_daily_plan(
        wip_file=args.wip_file,
        plan_date=args.plan_date,
        output_file=args.output,
        days=args.days,
        hide_empty_sections=args.hide_empty_sections,
    )

    return 0


# ---------------------------------------------------------------------------
# COLAB DIRECT RUN
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    example_args = [
        "--wip_file", WIP_FILE,
        "--plan_date", PLAN_DATE,
        "--output", OUTPUT_FILE,
        "--days", str(DAYS),
    ]

    sys.exit(main(example_args))
